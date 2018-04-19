import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import time
def get_url_list():
    MAXPAGE = 60
    idx = 1
    for pageIdx in range(1, MAXPAGE + 1):
        baseUrl = 'http://franchise.ftc.go.kr/user/extra/main/62/firMst/list/jsp/LayOutPage.do?column=&search=&searchFirRegNo=&selUpjong=&selIndus=&srow=1000&spage='
        url = baseUrl + str(pageIdx)
        html = requests.get(url)
        bs4 = BeautifulSoup(html.text, 'lxml')

        trs = bs4.find('div', id='txt').find('table').find_all('tr')[1:]
        for tr in trs:
            title = tr.find_all('td')[1].get_text().strip()
            link = tr.find_all('td')[1].a['href']
            print(idx, link)
            idx += 1
if __name__=="__main__":
    wb = Workbook()
    ws1 = wb.active
    ws1["A1"] = '상호명'
    ws1.column_dimensions["A"].width = 30
    ws1["B1"] = '전화'
    ws1.column_dimensions["B"].width = 30
    ws1["C1"] = '팩스'
    ws1.column_dimensions["C"].width = 30
    ws1["D1"] = '주소'
    ws1.column_dimensions["D"].width = 100
    ws1["E1"] = '가맹점현황'
    ws1.column_dimensions["E"].width = 10
    row = 2

    origin = 'http://franchise.ftc.go.kr'
    f = open('url.txt')
    lines = f.readlines()[5967:]
    idx = 1
    try:
        for line in lines:
            name = '' #상호
            tel = '' #전화번호
            fax = '' #팩스번호
            address = ''#주소
            cnt = ''#가맹점수
            while True:
                try:
                    url = origin+line.split(' ')[1].strip()
                    break
                except:
                    time.sleep(2)
            html = requests.get(url)
            bs4 = BeautifulSoup(html.text, 'lxml')
            div = bs4.find('div',id='txt')
            # 상호
            box1 = div.find('div',class_='box_flop')
            table1 = box1.find('table')
            name = table1.find_all('tr')[1].find('td')
            name.label.decompose()
            name = name.get_text().strip()
            #전화번호, 팩스
            tel = table1.find_all('tr')[3].find_all('td')[2].get_text().strip()
            fax = table1.find_all('tr')[3].find_all('td')[3].get_text().strip()
            # 주소
            table2 = box1.find_all('table')[1]
            address = table2.find('tr').find('td').get_text().strip()
            #가맹점수
            box2 = div.find_all('div', class_='box_flop')[1]
            table1 = box2.find_all('table')[1]
            cnt = table1.find('tbody').find('tr').find_all('td')[1].get_text().strip()
            ws1.cell(row=row,column=1,value=name)
            ws1.cell(row=row, column=2, value=tel)
            ws1.cell(row=row, column=3, value=fax)
            ws1.cell(row=row, column=4, value=address)
            ws1.cell(row=row, column=5, value=cnt)
            ws1.cell(row=row, column=6, value=url)
            row += 1
            print(idx)
            idx+=1
    except:
        wb.save('./data.xlsx')
    wb.save('./data.xlsx')

